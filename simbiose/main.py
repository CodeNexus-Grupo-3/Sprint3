from fastapi import FastAPI
from fastapi.responses import FileResponse
import mysql.connector
from datetime import date
import os
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

load_dotenv()
app = FastAPI()

'''
---> CONFIGURAÇÕES DE ESTILO <---
''' 

fill_header = PatternFill("solid", fgColor="ffd966")
fill_values = PatternFill("solid", fgColor="e2efda")

line = Side(style="thin", color="000000")
border = Border(left=line, right=line, top=line, bottom=line)

center_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")

# Variáveis de fonte
font_bold = Font(bold=True, size=13)
font_normal = Font(size=11)

'''
---> FUNÇÕES AUXILIARES <---
'''

def style_cell(cell, fill=None, bold=False, align=False):
    cell.border = border
    cell.fill = fill

    if bold:
        cell.font = font_bold
    else:
        cell.font = font_normal
    
    if align:
        cell.alignment = center_align
    else:
        cell.alignment = left_align

def write_cell(ws, row, col, value):
    return ws.cell(row, col, value=value)

def create_cell(ws, row, col, value, fill, isBold, isAlign):
    cell = write_cell(ws, row, col, value)
    style_cell(cell, fill, isBold, isAlign)

def resize(ws, col, row, size_row):
    col_letter = get_column_letter(col)
    ws.column_dimensions[col_letter].width = 22
    ws.row_dimensions[row].height = size_row

'''
---> FUNÇÃO PRINCIPAL <---
'''

def gerar_relatorio():

    '''
    ---> EXTRAINDO DADOS DO BD <---
    '''

    conexao = mysql.connector.connect(
        host=os.getenv("DB_HOST"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
        database=os.getenv("DB_NAME")
    )

    cursor = conexao.cursor(dictionary=True)

    cursor.execute("SELECT dtPartida, resultado, duracao, tipo, totalAbates, totalAssistencias, totalMortes, totalGold, totalBaron, totalDrag, totalTorres, totalDano FROM PartidasEquipe;")

    team_data = cursor.fetchall()

    cursor.close()
    conexao.close()

    '''
    ---> CRIAÇÃO DA PLANILHA <---
    '''

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio"

    header_titles = [
        "Data", 
        "Resultado", 
        "Duração",
        "Tipo", 
        "Total de Abates", 
        "Total de Assistências", 
        "Total de Mortes", 
        "Total de Gold", 
        "Total de Baron", 
        "Total de Drag", 
        "Total de Torres", 
        "Total de Dano"
    ]

    current_row = 2
    current_col = 2

    for i, titles in enumerate(header_titles, start=2):
        create_cell(ws, current_row, i, titles, fill_header, True, True)
        resize(ws, i, current_row, 27)

    for item in team_data:
        current_row += 1

        for key, value in item.items():

            if key == "resultado":
                value = "Vitória" if value == 1 else "Derrota"

            if key == "duracao":
                minutos = value // 60
                segundos = value % 60
                value = f"{minutos:02d}:{segundos:02d}"

            create_cell(ws, current_row, current_col, value, fill_values, False, False)
            resize(ws, current_col, current_row, 15)

            current_col += 1

        current_col = 2

    data = date.today().strftime("%d-%m-%Y")
    arquivo_saida = f"Relatorio-Partidas-Equipe-{data}.xlsx"
    wb.save(arquivo_saida)

    return arquivo_saida

'''
---> ROTA FASTAPI <---
'''

@app.post("/gerar-relatorio")
def executar_relatorio():

    caminho_arquivo = gerar_relatorio()

    return FileResponse(
        path=caminho_arquivo,
        filename="Relatorio.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )