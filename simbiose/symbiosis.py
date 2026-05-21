import mysql.connector
from datetime import date
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

'''
---> EXTRAINDO DADOS DO BD <---
'''

conexao = mysql.connector.connect(
    host="127.0.0.1",
    user="root",
    password="Polentinha69?",
    database="codenexus"
)

cursor = conexao.cursor(dictionary=True)

cursor.execute("SELECT dtPartida, resultado, duracao, tipo, totalAbates, totalAssistencias, totalMortes, totalGold, totalBaron, totalDrag, totalTorres, totalDano FROM PartidasEquipe;")

teamData = cursor.fetchall()

cursor.close()
conexao.close()

'''
---> CONFIGURAÇÕES DE ESTILO <---
''' 

wb = Workbook()
ws = wb.active
ws.title = "Relatorio"

fill_header = PatternFill("solid", fgColor="ffd966")
fill_values = PatternFill("solid", fgColor="e2efda")

line = Side(style="thin", color="000000")
border = Border(left=line, right=line, top=line, bottom=line)

center_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")

# Variáveis de fonte
font_bold = Font(bold=True, size=13)
font_normal = Font(size=11)

'''
---> FUNÇÕES AUXILIARES <---
'''

def style_cell(cell, fill=None, bold=False, align=False):
    cell.border = border
    cell.fill = fill

    if bold:
        cell.font = font_bold
    else:
        cell.font = font_normal
    
    if align:
        cell.alignment = center_align
    else:
        cell.alignment = left_align

def write_cell(row, col, value):
    return ws.cell(row, col, value=value)

def create_cell(row, col, value, fill, isBold, isAlign):
    cell = write_cell(row, col, value)
    style_cell(cell, fill, isBold, isAlign)

def resize(col, row, size_row):
    col_letter = get_column_letter(col)
    ws.column_dimensions[col_letter].width = 22
    ws.row_dimensions[row].height = size_row

header_titles = [
    "Data", 
    "Resultado", 
    "Duração",
    "Tipo", 
    "Total de Abates", 
    "Total de Assistências", 
    "Total de Mortes", 
    "Total de Gold", 
    "Total de Baron", 
    "Total de Drag", 
    "Total de Torres", 
    "Total de Dano"
]

current_row = 2
current_col = 2

for i, titles in enumerate(header_titles, start=2):
    create_cell(current_row, i, titles, fill_header, True, True)
    resize(i, current_row, 27)

for item in teamData:
    current_row += 1

    for key, value in item.items():

        if key == "resultado":
            value = "Vitória" if value == 1 else "Derrota"

        if key == "duracao":
            minutos = value // 60
            segundos = value % 60
            value = f"{minutos:02d}:{segundos:02d}"

        create_cell(current_row, current_col, value, fill_values, False, False)
        resize(current_col, current_row, 15)
        print(value)

        current_col += 1

    current_col = 2

data = date.today().strftime("%d-%m-%Y")
downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
arquivo_saida = os.path.join(downloads_path, f"Relatorio-Partidas-Equipe-{data}.xlsx")
wb.save(arquivo_saida)

print("Arquivo gerado com sucesso: " + arquivo_saida)